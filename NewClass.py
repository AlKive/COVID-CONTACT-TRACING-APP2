#PSEUDOCODE
#import tkinter 
from tkinter import*
from tkinter import ttk
import tkinter
import os
import openpyxl
from tkinter import messagebox




#create class
class Tracer:
    #init function & create instance of variable
    def __init__(self):
            
            #create tkinter window   
            self.ws = tkinter.Tk()
            self.ws.geometry("1500x1500")
            self.ws.title("COVID Contact Tracing Form")
            self.ws.configure(highlightbackground= "dark blue", highlightthickness= 6)

        #create frame for necessary information 
            self.user_info_frame = tkinter.Frame(self.ws, background= "light blue", highlightthickness= 4, highlightbackground= "dark blue")
            self.user_info_frame.place(x= 10, y = 20, width = 350, height = 600)
            #label the frame
            self.user_info_frame_label =tkinter.LabelFrame(self.user_info_frame, text="PERSONAL INFORMATION",
                                                           padx=5, pady= 20, font= ("Brass Mono", 13),
                                                            highlightthickness= 4, highlightbackground= "light yellow", 
                                                            foreground= "dark blue", highlightcolor= "dark green")
            self.user_info_frame_label.place(x= 10, y = 20, width= 300, height= 500)

            #create frame for displaying symptoms
            self.display_symptoms_frame = tkinter.Frame(self.ws, background= "light blue", highlightthickness= 4, highlightbackground= "dark blue")
            self.display_symptoms_frame.place(x= 800, y = 20, width = 180, height = 450)
            #label the frame
            self.display_symptoms_frame_label =tkinter.LabelFrame(self.display_symptoms_frame, text= "SYMPTOMS",
                                                           padx=5, pady= 20, font= ("Brass Mono", 13),
                                                            highlightthickness= 4, highlightbackground= "light yellow", 
                                                            foreground= "dark blue", highlightcolor= "dark green")
            self.display_symptoms_frame_label.place(x= 10, y = 20, width= 160, height= 420)

        #create frame for checkboxes
            self.symptoms_frame= tkinter.Frame(self.ws, background= "light blue",highlightthickness= 4, highlightbackground= "dark blue")
            self.symptoms_frame.place (x= 400, y = 20, width = 380, height = 600)
            #label the frame
            self.symptoms_frame_label =tkinter.LabelFrame(self.symptoms_frame, text=  "HEALTH INFORMATION ", 
                                                          padx=5, pady= 20, font= ("Brass Mono", 13),
                                                          foreground= "dark blue", highlightbackground ="light yellow", 
                                                          highlightthickness= 4, highlightcolor= "dark green")
            self.symptoms_frame_label.place(x= 10, y = 20, width= 360, height= 500)
            
    #create functions for getting user's input using tkinter entry with label

        #first name label and entry
    def firstName(self):
        self.FIRST_NAME = tkinter.Label(self.user_info_frame, text="First Name", highlightbackground= "dark blue", highlightthickness= 2)
        self.FIRST_NAME.place(x= 25, y = 70)
        self.FIRST_NAME_INPUT = tkinter.Entry(self.user_info_frame)
        self.FIRST_NAME_INPUT.place(x= 25, y = 95)
        
        #last name label and entry
    def lastName(self):
        self.LAST_NAME = tkinter.Label(self.user_info_frame, text="Last Name", highlightbackground= "dark blue", highlightthickness= 2)
        self.LAST_NAME.place(x= 25, y = 125)
        self.LAST_NAME_INPUT = tkinter.Entry(self.user_info_frame)
        self.LAST_NAME_INPUT.place(x= 25, y = 150)

    #Gender label and entry
    def gender(self):
        self.GENDER = tkinter.Label(self.user_info_frame, text="Gender", highlightbackground= "dark blue", highlightthickness= 2)
        self.GENDER.place(x= 25, y = 180)
        self.GENDER_INPUT = ttk.Combobox(self.user_info_frame,values=["Male", "Female", "Transgender Male", "Transgender Female","Gender Neutral", "non-binary", "agender"])
        self.GENDER_INPUT.place(x= 25, y = 205)

    
    #Age label and entry
    def age(self):
        self.AGE = tkinter.Label(self.user_info_frame, text="Age", highlightbackground= "dark blue", highlightthickness= 2)
        self.AGE.place(x= 25, y = 235)
        self.AGE_INPUT = tkinter.Spinbox(self.user_info_frame, from_=1, to=100)
        self.AGE_INPUT.place(x= 25, y = 260)

        
    #Vaccination Status label and entry
    def vaccination(self):
        #"Unvaccinated", "1st Dose", "2nd Dose(Fully Vaccinated)", "1st Booster Shot", "2nd Booster Shot"
        self.VACCINATION_STATUS = tkinter.Label(self.symptoms_frame, text="Vaccination Status", highlightbackground= "dark blue", highlightthickness= 2)
        self.VACCINATION_STATUS.place(x= 40, y = 250)
        self.VACCINATION_STATUS_INPUT = ttk.Combobox(self.symptoms_frame, values= ["Unvaccinated", "1st Dose", "2nd Dose(Fully Vaccinated)", "1st Booster Shot", "2nd Booster Shot"])
        self.VACCINATION_STATUS_INPUT['state'] = 'readonly'
        self.VACCINATION_STATUS_INPUT.place(x= 40, y = 275)

    #Phone Number label and entry
    def phone(self):
        self.PHONE = tkinter.Label(self.user_info_frame, text= "Cellphone Number", highlightbackground= "dark blue", highlightthickness= 2)
        self.PHONE.place(x= 25, y = 300)
        self.PHONE_INPUT = tkinter.Entry(self.user_info_frame)
        self.PHONE_INPUT.place(x= 25, y = 325)
    # address label and entry
    def address(self):
        self.ADDRESS = tkinter.Label(self.user_info_frame, text="Address", highlightbackground= "dark blue", highlightthickness= 2)
        self.ADDRESS.place(x= 25, y = 355)
        self.ADDRESS_INPUT = tkinter.Entry(self.user_info_frame)
        self.ADDRESS_INPUT.place(x= 25, y = 380)
        
    def rel(self):
        self.RELATIONSHIP = tkinter.Label(self.user_info_frame, text="Relationship to Contact Person", highlightbackground= "dark blue", highlightthickness= 2)
        self.RELATIONSHIP.place(x= 25, y = 410)
        self.RELATIONSHIP_INPUT = ttk.Combobox(self.user_info_frame, values= ["Parent", "Friend", "Co-Worker", "Schoolmate", "Partner", "Customer"])
        self.RELATIONSHIP_INPUT.place(x= 25, y = 435)       
    #put terms and condition checkbox
    def accept(self):
        self.accept_var = tkinter.StringVar(value="Not Accepted")
        self.terms_check = tkinter.Checkbutton(self.user_info_frame, text= "I accept the terms and conditions.",
                                        variable=self.accept_var, onvalue="Accepted", offvalue="Not Accepted")
        self.terms_check.place(x = 25, y = 470)
    #fix spacing of widgets
    
            
            
            
            
    '''###############################################################################################################################################################
   #########################################################################################################################'''
    def question_label(self):
        self.Q1 =Label(self.symptoms_frame, text = "1. Are you experiencing any symptoms such as:", fg = "dark blue", bg = "white", font=("Brass Mono",10))
        self.Q1.place(x =30 , y = 50)
        
        self.Q2 =Label(self.symptoms_frame, text = "2. Have you been vaccinated for COVID-19?", fg = "dark blue", bg = "white", font=("Brass Mono",10))
        self.Q2.place(x =30 , y = 220)
        
        self.Q3 =Label(self.symptoms_frame, text = "3.  Have you been exposed to a probable or confirmed ", fg = "dark blue", bg = "white", font=("Brass Mono",10))
        self.Q3.place(x =30 , y = 300)
        self.Q3_2 =Label(self.symptoms_frame, text = "   case  in the last 14 days?", fg = "dark blue", bg = "white", font=("Brass Mono",10))
        self.Q3_2.place(x =30 , y = 325)
    
    def symptoms(self):
        #FEVER
        self.fever = tkinter.StringVar(value= "None")
        self.fever_box = tkinter.Checkbutton(self.symptoms_frame, text= "fever", variable=self.fever, onvalue="Fever", offvalue="None")
        self.fever_box.place(x=40, y=70)       
        
        #COUGH
        self.cough = tkinter.StringVar(value= "None")
        self.cough_box= tkinter.Checkbutton(self.symptoms_frame, text= "cough", variable=self.cough, onvalue="cough", offvalue="None")
        self.cough_box.place(x=40, y= 90)     
        
        #COLD
        self.cold = tkinter.StringVar(value= "None")
        self.cold_box= tkinter.Checkbutton(self.symptoms_frame, text= "cold", variable=self.cold, onvalue="cold", offvalue="None")
        self.cold_box.place(x=40, y= 110)  
        
        #SORE THROAT
        self.sore_throat = tkinter.StringVar(value= "None")
        self.sore_throat_box= tkinter.Checkbutton(self.symptoms_frame, text= "sore throat", variable=self.sore_throat, onvalue="sore throat", offvalue="None")
        self.sore_throat_box.place(x=40, y= 130)  
        
        #MUSCLE PAIN
        self.muscle_pain = tkinter.StringVar(value= "None")
        self.muscle_pain_box= tkinter.Checkbutton(self.symptoms_frame, text= "muscle pain", variable=self.muscle_pain, onvalue="muscle pain", offvalue="None")
        self.muscle_pain_box.place(x=40, y= 160)  
        
        #HEADACHE
        self.headache = tkinter.StringVar(value= "None")
        self.headache_box= tkinter.Checkbutton(self.symptoms_frame, text= "headache", variable=self.headache, onvalue="headache", offvalue="None")
        self.headache_box.place(x=200, y= 70)  
        
        #FATIGUE
        self.fatigue = tkinter.StringVar(value= "None")
        self.fatigue_box= tkinter.Checkbutton(self.symptoms_frame, text= "fatigue", variable=self.fatigue, onvalue="fatigue", offvalue="None")
        self.fatigue_box.place(x=200, y=90 )  
        
        #SHORTNESS OF BREATH
        self.shortness_breath = tkinter.StringVar(value= "None")
        self.shortness_breath_box= tkinter.Checkbutton(self.symptoms_frame, text= "shortness of Breath", variable=self.shortness_breath, onvalue="shortness of Breath", offvalue="None")
        self.shortness_breath_box.place(x=200, y= 110)  
        
        #LOSS OF SMELL
        self.loss_smell = tkinter.StringVar(value= "None")
        self.loss_smell_box= tkinter.Checkbutton(self.symptoms_frame, text= "loss of smell", variable=self.loss_smell, onvalue="loss of smell", offvalue="None")
        self.loss_smell_box.place(x=200, y=130 ) 
        
        #LOSS OF TASTE
        self.loss_taste = tkinter.StringVar(value= "None")
        self.loss_taste_box= tkinter.Checkbutton(self.symptoms_frame, text= "loss of taste", variable=self.loss_taste, onvalue="loss of taste", offvalue="None")
        self.loss_taste_box.place(x=200, y=160 ) 
        
        #NONE OF THE ABOVE
        self.NOTA = tkinter.StringVar(value="some symptoms acquired")
        self.NOTA_box= tkinter.Checkbutton(self.symptoms_frame, text= "None of the above", variable=self.NOTA, onvalue="None", offvalue = 'some symptoms acquired')
        self.NOTA_box.place(x=120, y= 190) 
       
       
    def display_symp(self):
        #FEVER
        self.fever_entry = tkinter.Entry(self.display_symptoms_frame)
        self.fever_entry.place(x=25, y=90)       
        
        #COUGH
        self.cough_entry= tkinter.Entry(self.display_symptoms_frame)
        self.cough_entry.place(x=25, y= 120)     
        
        #COLD
      
        self.cold_entry= tkinter.Entry(self.display_symptoms_frame)
        self.cold_entry.place(x=25, y= 150)  
        
        #SORE THROAT
        self.sore_throat_entry= tkinter.Entry(self.display_symptoms_frame)
        self.sore_throat_entry.place(x=25, y=180)  
        
        #MUSCLE PAIN
        self.muscle_pain_entry= tkinter.Entry(self.display_symptoms_frame)
        self.muscle_pain_entry.place(x=25, y= 210)  
        
        #HEADACHE
        self.headache_entry= tkinter.Entry(self.display_symptoms_frame)
        self.headache_entry.place(x=25, y= 240)  
        
        #FATIGUE
        self.fatigue_entry= tkinter.Entry(self.display_symptoms_frame)
        self.fatigue_entry.place(x=25, y=270 )  
        
        #SHORTNESS OF BREATH
        self.shortness_breath_entry= tkinter.Entry(self.display_symptoms_frame)
        self.shortness_breath_entry.place(x=25, y= 300)  
        
        #LOSS OF SMELL
        self.loss_smell_entry= tkinter.Entry(self.display_symptoms_frame)
        self.loss_smell_entry.place(x=25, y=330 ) 
        
        #LOSS OF TASTE
        self.loss_taste_entry= tkinter.Entry(self.display_symptoms_frame)
        self.loss_taste_entry.place(x=25, y=360 ) 
        
        #NONE OF THE ABOVE
        self.NOTA_entry= tkinter.Entry(self.display_symptoms_frame)
        self.NOTA_entry.place(x=300, y= 390)   
    
    #Nationality label and entry
    def exposure(self):
        #Nationality label and entry
        self.EXPOSURE = tkinter.Label(self.symptoms_frame, text="Exposure", highlightbackground= "dark blue", highlightthickness= 2)
        self.EXPOSURE.place(x= 40, y = 360)
        self.EXPOSURE_INPUT = ttk.Combobox(self.symptoms_frame, values=["Yes", "No", "Uncertain"])
        self.EXPOSURE_INPUT['state'] = 'normal'
        self.EXPOSURE_INPUT.place(x= 40, y = 385)    
        
                                             





    #create a save/add function for all user's info into a database ( i prefer excel as my database)
    def add(self):
        accepted = self.accept_var.get()
        nota = self.NOTA.get()        
        firstname = self.FIRST_NAME_INPUT.get()
        lastname = self.LAST_NAME_INPUT.get()
        exposure  = (self.EXPOSURE_INPUT).get() 
        age = self.AGE_INPUT.get()
        address = self.ADDRESS_INPUT.get()
        gender = self.GENDER_INPUT.get()
        vaccination = self.VACCINATION_STATUS_INPUT.get()
        phone = self.PHONE_INPUT.get()
        relationsip = self.RELATIONSHIP_INPUT.get()
        
        #get symptoms inputs
        fever = self.fever.get()
        cough=self.cough.get()
        cold =self.cold.get()
        headache =self.headache.get()
        sore_throat=self.sore_throat.get()
        muscle_pain = self.muscle_pain.get()
        shortness_of_breath= self.shortness_breath.get()
        fatigue= self.fatigue.get()
        loss_of_taste= self.loss_taste.get()
        loss_of_smell= self.loss_smell.get()
        
        if accepted=="Accepted":
            if nota == "some symptoms acquired" and nota != "None":
                if  firstname and lastname and age and gender and exposure and address and phone and vaccination and relationsip and fever and cough and cold and sore_throat and headache and muscle_pain  and fatigue and shortness_of_breath and loss_of_smell and loss_of_taste :

                    
            
                    filepath =("C:\git\BSCPE 1ST YEAR 2ND SEM\OOP\ASSIGNMENTS\COVID-CONTACT-TRACING-APP\Tracer\data.xlsx")

                    if not os.path.exists(filepath):
                        workbook = openpyxl.Workbook()
                        sheet = workbook.active
                        heading = ["First Name", "Last Name", "Gender", "Age", "Nationality", "Address", "Cellphone Number", 
                                   "Vaccination status", "Relationship to Contact Person", "fever", "cough", "cold",
                                   "sore_throat", "headache", "muscle_pain", "fatigue", "shortness_of_breath",
                                   "loss_of_smell", "loss_of_taste"]
                        sheet.append(heading)
                        workbook.save(filepath)
                    workbook = openpyxl.load_workbook(filepath)
                    sheet = workbook.active
                    sheet.append([firstname, lastname, gender, age, exposure, address,
                                phone, vaccination, relationsip,fever, cough, cold,
                                   sore_throat, headache, muscle_pain, fatigue, shortness_of_breath,
                                   loss_of_smell, loss_of_taste])
                    workbook.save(filepath)
                    messagebox.showinfo("DATA", "DATA SAVED")
                
                else:
                    messagebox.showwarning(title="Error", message="All information are required.")
                    
            elif  nota == "None" and nota != "some symptoms acquired":
                self.fever_box.invoke()
                self.cough_box.invoke()
                self.cold_box.invoke()
                self.headache_box.invoke()
                self.sore_throat_box.invoke()
                self.muscle_pain_box.invoke()
                self.shortness_breath_box.invoke()
                self.fatigue_box.invoke()
                self.loss_taste_box.invoke()
                self.loss_smell_box.invoke()
                
                
                firstname = self.FIRST_NAME_INPUT.get()
                lastname = self.LAST_NAME_INPUT.get()
                exposure  = (self.EXPOSURE_INPUT).get() 
                age = self.AGE_INPUT.get()
                address = self.ADDRESS_INPUT.get()
                gender = self.GENDER_INPUT.get()
                vaccination = self.VACCINATION_STATUS_INPUT.get()
                phone = self.PHONE_INPUT.get()
                relationsip = self.RELATIONSHIP_INPUT.get()
                
                if  firstname and lastname and age and gender and exposure and address and phone and vaccination and relationsip and nota :
                    filepath =("C:\git\BSCPE 1ST YEAR 2ND SEM\OOP\ASSIGNMENTS\COVID-CONTACT-TRACING-APP\Tracer\data.xlsx")
                    if not os.path.exists(filepath):
                        workbook = openpyxl.Workbook()
                        sheet = workbook.active
                        heading = ["First Name", "Last Name", "Gender", "Age", "Exposure", "Address", "Cellphone Number", 
                                   "Vaccination status", "Relationship to Contact Person", "fever", "cough", "cold",
                                   "sore_throat", "headache", "muscle_pain", "fatigue", "shortness_of_breath",
                                   "loss_of_smell", "loss_of_taste"]
                        sheet.append(heading)
                        workbook.save(filepath)
                    workbook = openpyxl.load_workbook(filepath)
                    sheet = workbook.active
                    sheet.append([firstname, lastname, gender, age, exposure, address,
                                phone, vaccination, relationsip, fever, cough, cold,
                                   sore_throat, headache, muscle_pain, fatigue, shortness_of_breath,
                                   loss_of_smell, loss_of_taste] )
                    workbook.save(filepath)
                    messagebox.showinfo("DATA", "DATA SAVED")
                    
                
                else:
                    messagebox.showwarning(title= "Error", message="All information are required.")
            else:
                if nota != "None" and nota != "some symptoms acquired":
                    messagebox.showerror('ERROR', 'Select your symptoms or Select None.')      
        else:
            messagebox.showerror("ERROR", "You have not accepted the terms and conditions")        
                
    #create search method
    def searchByFirstName(self):
        FirstNameInput = self.FIRST_NAME_INPUT.get()
        self.FIRST_NAME_INPUT.configure(state= tkinter.NORMAL)
        self.LAST_NAME_INPUT.configure(state= tkinter.NORMAL)
        self.GENDER_INPUT.configure(state= tkinter.NORMAL)
        self.AGE_INPUT.configure(state= tkinter.NORMAL)
        self.EXPOSURE_INPUT.configure(state= tkinter.NORMAL)
        self.ADDRESS_INPUT.configure(state= tkinter.NORMAL)
        self.PHONE_INPUT.configure(state= tkinter.NORMAL)
        self.VACCINATION_STATUS_INPUT.configure(state= tkinter.NORMAL)
        self.RELATIONSHIP_INPUT.configure(state= tkinter.NORMAL)
        
        
    
        
        self.FIRST_NAME_INPUT.delete(0, 'end')
        self.LAST_NAME_INPUT.delete(0, 'end')
        self.GENDER_INPUT.delete(0, 'end')
        self.AGE_INPUT.delete(0, 'end')
        self.EXPOSURE_INPUT.delete(0, 'end')
        self.ADDRESS_INPUT.delete(0, 'end')
        self.PHONE_INPUT.delete(0, 'end')
        self.VACCINATION_STATUS_INPUT.delete(0, 'end')
        self.RELATIONSHIP_INPUT.delete(0,'end')
        
        
        
        
        wb = openpyxl.load_workbook("C:\git\BSCPE 1ST YEAR 2ND SEM\OOP\ASSIGNMENTS\COVID-CONTACT-TRACING-APP\Tracer\data.xlsx")  
        sheet = wb["Sheet"]  
        for cell in sheet.iter_rows(min_row=1, min_col=1, max_row =sheet.max_row, max_col=19, values_only=True):  
            if cell[0] == str(FirstNameInput):
                self.FIRST_NAME_INPUT.insert(0, cell[0])
                self.LAST_NAME_INPUT.insert(0, cell[1])
                self.GENDER_INPUT.insert(0, cell[2])
                self.AGE_INPUT.insert(0, cell[3])
                self.EXPOSURE_INPUT.insert(0, cell[4])
                self.ADDRESS_INPUT.insert(0, cell[5])
                self.PHONE_INPUT.insert(0, cell[6])
                self.VACCINATION_STATUS_INPUT.insert(0, cell[7])
                self.RELATIONSHIP_INPUT.insert(0, cell[8])
                self.fever_entry.insert(0, cell[9])
                self.cough_entry.insert(0, cell[10])
                self.cold_entry .insert(0, cell[11])
                self.headache_entry .insert(0, cell[12])
                self.sore_throat_entry.insert(0, cell[13])
                self.muscle_pain_entry.insert(0, cell[14])
                self.shortness_breath_entry.insert(0, cell[15])
                self.fatigue_entry.insert(0, cell[16])
                self.loss_taste_entry.insert(0, cell[17])
                self.loss_smell_entry.insert(0, cell[18])
                
        
                
                self.FIRST_NAME_INPUT.configure(state= tkinter.NORMAL)
                self.LAST_NAME_INPUT.configure(state= tkinter.NORMAL)
                self.GENDER_INPUT.configure(state= tkinter.NORMAL)
                self.AGE_INPUT.configure(state= tkinter.NORMAL)
                self.EXPOSURE_INPUT.configure(state= tkinter.NORMAL)
                self.ADDRESS_INPUT.configure(state= tkinter.NORMAL)
                self.PHONE_INPUT.configure(state= tkinter.NORMAL)
                self.VACCINATION_STATUS_INPUT.configure(state= tkinter.NORMAL)
                self.RELATIONSHIP_INPUT.configure(state= tkinter.NORMAL)
        
                
                
                
                
    def clear(self):
            self.FIRST_NAME_INPUT.delete(0, "end")
            self.LAST_NAME_INPUT.delete(0, 'end')
            self.GENDER_INPUT.delete(0, 'end')
            self.AGE_INPUT.delete(0, 'end')
            self.EXPOSURE_INPUT.delete(0, 'end')
            self.ADDRESS_INPUT.delete(0, 'end')
            self.PHONE_INPUT.delete(0, 'end')
            self.VACCINATION_STATUS_INPUT.set('0')
            self.RELATIONSHIP_INPUT.delete(0,'end')
            self.fever_box.deselect()
            self.sore_throat_box.deselect()
            self.cold_box.deselect()
            self.cough_box.deselect()
            self.headache_box.deselect()
            self.muscle_pain_box.deselect()
            self.fatigue_box.deselect()
            self.shortness_breath_box.deselect()
            self.loss_smell_box.deselect()
            self.loss_taste_box.deselect()
            self.NOTA_box.deselect()
            self.fever_entry.delete(0, "end")
            self.cough_entry.delete(0, "end")
            self.cold_entry.delete(0, "end")
            self.headache_entry.delete(0, "end")
            self.sore_throat_entry.delete(0, "end")
            self.muscle_pain_entry.delete(0, "end")
            self.shortness_breath_entry.delete(0, "end")
            self.fatigue_entry.delete(0, "end")
            self.loss_taste_entry.delete(0, "end")
            self.loss_smell_entry.delete(0, "end")
            self.NOTA_entry.delete(0, "end")
                        
    #create save/add button 
    def button_add(self):
        button = tkinter.Button(self.symptoms_frame, text="SAVE", command= self.add)
        button.place(x= 260, y= 450)    
            
    #create search button
    def button_search(self):
        button = tkinter.Button(self.symptoms_frame, text="SEARCH", command= self.searchByFirstName)
        button.place(x= 60, y= 450)
        
    def button_clear(self):
        button = tkinter.Button(self.symptoms_frame, text="CLEAR", command= self.clear)
        button.place(x= 160, y= 450)    
        